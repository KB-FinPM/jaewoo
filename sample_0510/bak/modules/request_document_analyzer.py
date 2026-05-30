import os
import json
import boto3

from dotenv import load_dotenv
from datetime import datetime
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def save_to_excel(content, timestamp):
    """
    최종 분석 결과를 엑셀 파일에 요구사항 명세서 표 형태로 저장
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "분석 결과"

    # 스타일 정의
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    center_alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
    left_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    def split_row(line: str):
        if "|" in line:
            parts = [part.strip() for part in line.strip().split("|")]
            if parts and parts[0] == "":
                parts = parts[1:]
            if parts and parts[-1] == "":
                parts = parts[:-1]
            return parts
        if "\t" in line:
            return [part.strip() for part in line.split("\t")]
        if "," in line:
            return [part.strip() for part in line.split(",")]
        return [line.strip()]

    def is_separator_row(line: str):
        return bool(re.fullmatch(r"[\s|:\-]+", line))

    def extract_table(lines):
        headers = ["업무", "구분", "요구사항ID", "요구사항명", "기능/비기능 요구사항", "비고"]
        for idx, line in enumerate(lines):
            if "|" in line or "," in line or "\t" in line:
                row_values = split_row(line)
                matched = sum(1 for v in row_values if v in headers)
                if matched >= 3:
                    table_header = row_values
                    data_lines = lines[idx+1:]
                    if data_lines and is_separator_row(data_lines[0]):
                        data_lines = data_lines[1:]
                    table_rows = []
                    for data_line in data_lines:
                        row_values = split_row(data_line)
                        if not row_values:
                            continue
                        if len(row_values) < len(table_header):
                            row_values += [""] * (len(table_header) - len(row_values))
                        table_rows.append(row_values[:len(table_header)])
                    return table_header, table_rows
        return [], []

    lines = [line.rstrip() for line in content.strip().splitlines() if line.strip()]
    headers = ["업무", "구분", "요구사항ID", "요구사항명", "기능/비기능 요구사항", "비고"]
    row = 2

    table_header, table_rows = extract_table(lines)

    if table_header and table_rows:
        for col, header in enumerate(table_header, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
        row += 1

        for row_data in table_rows:
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.alignment = left_alignment
                cell.border = border
            row += 1
    else:
        # 테이블 형식이 아니면 기본 요약 형식으로 저장
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
        row += 1

        for line in lines:
            cell = ws.cell(row=row, column=1)
            cell.value = "요구사항"
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            cell.alignment = left_alignment
            cell.border = border

            content_cell = ws.cell(row=row, column=2)
            content_cell.value = line
            content_cell.alignment = left_alignment
            content_cell.border = border
            row += 1

    widths = [20, 15, 20, 40, 30, 20]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[chr(ord("A") + idx - 1)].width = width

    for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
        for cell in row_cells:
            if cell.value is not None:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    output_dir = "./results"
    os.makedirs(output_dir, exist_ok=True)
    filename = timestamp.replace(":", "").replace("-", "").replace(" ", "_")
    excel_path = os.path.join(output_dir, f"analysis_{filename}.xlsx")
    wb.save(excel_path)
    return excel_path


def request_document_analyzer(chunk_summaries):
    # 3. bedrock client 설정
    load_dotenv()
    model_id = os.getenv("MODEL_ID")
    client = boto3.client("bedrock-runtime", region_name="ap-northeast-2")

    # 토큰 사용량 요약
    token_summaries = [ 0 , 0 ]   # [입력 토큰 총합, 출력 토큰 총합]

    # 6. 통합 분석
    combined_summary = "\n".join(chunk_summaries)

    print("통합 분석 데이터 시작 =====================================")
    print(combined_summary)
    print("통합 분석 데이터 끝 =====================================")

    final_request_body = {
        "anthropic_version": "bedrock-2023-05-31",
        "max_tokens": 2000,
        "messages": [
            {
                "role": "user",
                "content": f"""
                    당신은 프로젝트 매니저(PM)입니다.
                    아래는 문서의 여러 부분을 분석한 결과입니다.
                    이들을 종합하여 요구사항 명세서를 만드려고 합니다.

                    분석 결과:
                    {combined_summary}

                    최종 정리:
                    요구사항 명세서 엑셀로 저장할 수 있도록 다음 컬럼을 가진 표 형식으로 정리해주세요:
                    업무, 구분, 요구사항ID, 요구사항명, 기능/비기능 요구사항, 비고
                    각 요구사항은 한 줄에 하나씩 작성하고, 출력은 Markdown 또는 CSV 표 형태로 해주세요.
                    """
            }
        ]
    }

    final_response = client.invoke_model(
        modelId=model_id,
        body=json.dumps(final_request_body)
    )
    final_response_body = json.loads(final_response["body"].read())

    print("통합 분석 완료 =====================================")

    # 토큰 사용량 업데이트
    usage = final_response_body.get("usage", {})
    token_summaries[0] += usage.get("input_tokens", 0)
    token_summaries[1] += usage.get("output_tokens", 0) 

    # 최종 결과 텍스트 추출
    final_text = final_response_body["content"][0]["text"]

    # 7. 최종 결과 출력
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f" \n Request_Document_Analyzer - 입력 토큰, 출력 토큰, 토큰 합계 : {token_summaries [ 0 ]} / {token_summaries [ 1 ]} / {token_summaries [ 0 ] + token_summaries [ 1 ]} ")
    print(f"\n[{now}] 최종 분석 결과:\n{final_text}")

    # 8. 엑셀 파일에 저장
    excel_path = save_to_excel(final_text, now)
    print(f"\n[{now}] 엑셀 파일 저장 완료: {excel_path}")

    return final_text