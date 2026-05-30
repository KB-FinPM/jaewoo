import os
import re
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

EXCEL_HEADERS = [
    "업무",
    "구분",
    "요구사항ID",
    "요구사항명",
    "기능/비기능 요구사항",
    "비고",
]

EXCEL_COLUMN_WIDTHS = [20, 15, 20, 40, 30, 20]
TABLE_HEADER_THRESHOLD = 3
TABLE_DELIMITERS = ("|", "\t", ",")

STYLE_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
STYLE_HEADER_FONT = Font(bold=True, color="FFFFFF", size=12)
STYLE_CELL_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
STYLE_CENTER = Alignment(horizontal="center", vertical="top", wrap_text=True)
STYLE_LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)
STYLE_SUMMARY_FILL = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
STYLE_CHUNK_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")


def normalize_row(line: str) -> list[str]:
    if "|" in line:
        parts = [part.strip() for part in line.strip().split("|")]
        if parts and parts[0] == "":
            parts = parts[1:]
        if parts and parts[-1] == "":
            parts = parts[:-1]
        return parts
    if "\t" in line:
        return [part.strip() for part in line.split("\t")]
    if "," in line:
        return [part.strip() for part in line.split(",")]
    return [line.strip()]


def is_separator_row(line: str) -> bool:
    return bool(re.fullmatch(r"[\s|:\-]+", line))


def parse_table(lines: list[str]) -> tuple[list[str], list[list[str]]]:
    for idx, line in enumerate(lines):
        if any(delimiter in line for delimiter in TABLE_DELIMITERS):
            header_row = normalize_row(line)
            matched_headers = sum(1 for value in header_row if value in EXCEL_HEADERS)
            if matched_headers >= TABLE_HEADER_THRESHOLD:
                data_lines = lines[idx + 1 :]
                if data_lines and is_separator_row(data_lines[0]):
                    data_lines = data_lines[1:]

                table_rows: list[list[str]] = []
                for data_line in data_lines:
                    row_values = normalize_row(data_line)
                    if not any(value for value in row_values):
                        continue
                    if len(row_values) < len(header_row):
                        row_values += [""] * (len(header_row) - len(row_values))
                    table_rows.append(row_values[: len(header_row)])

                return header_row, table_rows

    return [], []


def create_workbook(excel_path: str):
    os.makedirs(os.path.dirname(excel_path), exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "분석 결과"

    for col, header in enumerate(EXCEL_HEADERS, start=1):
        cell = ws.cell(row=2, column=col)
        cell.value = header
        cell.font = STYLE_HEADER_FONT
        cell.fill = STYLE_HEADER_FILL
        cell.alignment = STYLE_CENTER
        cell.border = STYLE_CELL_BORDER

    for col_index, width in enumerate(EXCEL_COLUMN_WIDTHS, start=1):
        ws.column_dimensions[chr(ord("A") + col_index - 1)].width = width

    wb.save(excel_path)
    return wb, ws


def find_column_index(ws, header_name: str) -> Optional[int]:
    for col in range(1, len(EXCEL_HEADERS) + 1):
        if ws.cell(row=2, column=col).value == header_name:
            return col
    return None


def get_next_requirement_id(ws) -> int:
    id_col = find_column_index(ws, "요구사항ID")
    if id_col is None:
        return 1

    max_id = 0
    for row in range(3, ws.max_row + 1):
        value = ws.cell(row=row, column=id_col).value
        if isinstance(value, str) and value.strip().isdigit():
            max_id = max(max_id, int(value.strip()))
        elif isinstance(value, int):
            max_id = max(max_id, value)

    return max_id + 1


def normalize_request_id(value: str, next_id: int) -> tuple[str, int]:
    if not value:
        return str(next_id), next_id + 1

    stripped = value.strip()
    if stripped.isdigit():
        return str(next_id), next_id + 1

    if stripped.startswith("REQ-") and stripped[4:].isdigit():
        return f"REQ-{next_id}", next_id + 1

    return value, next_id


def save_to_excel(content: str, timestamp: str) -> str:
    lines = [line.rstrip() for line in content.strip().splitlines() if line.strip()]
    table_header, table_rows = parse_table(lines)

    output_dir = "./results"
    os.makedirs(output_dir, exist_ok=True)
    filename = timestamp.replace(":", "").replace("-", "").replace(" ", "_")
    excel_path = os.path.join(output_dir, f"analysis_{filename}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "분석 결과"

    if table_header and table_rows:
        for col, header in enumerate(table_header, start=1):
            cell = ws.cell(row=2, column=col)
            cell.value = header
            cell.font = STYLE_HEADER_FONT
            cell.fill = STYLE_HEADER_FILL
            cell.alignment = STYLE_CENTER
            cell.border = STYLE_CELL_BORDER

        for row_index, row_values in enumerate(table_rows, start=3):
            for col, value in enumerate(row_values, start=1):
                cell = ws.cell(row=row_index, column=col)
                cell.value = value
                cell.alignment = STYLE_LEFT
                cell.border = STYLE_CELL_BORDER
    else:
        for col, header in enumerate(EXCEL_HEADERS, start=1):
            cell = ws.cell(row=2, column=col)
            cell.value = header
            cell.font = STYLE_HEADER_FONT
            cell.fill = STYLE_HEADER_FILL
            cell.alignment = STYLE_CENTER
            cell.border = STYLE_CELL_BORDER

        for row_index, line in enumerate(lines, start=3):
            first = ws.cell(row=row_index, column=1)
            first.value = "요구사항"
            first.font = Font(bold=True, size=11)
            first.fill = STYLE_SUMMARY_FILL
            first.alignment = STYLE_LEFT
            first.border = STYLE_CELL_BORDER

            content_cell = ws.cell(row=row_index, column=2)
            content_cell.value = line
            content_cell.alignment = STYLE_LEFT
            content_cell.border = STYLE_CELL_BORDER

    for col_index, width in enumerate(EXCEL_COLUMN_WIDTHS, start=1):
        ws.column_dimensions[chr(ord("A") + col_index - 1)].width = width

    for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(EXCEL_HEADERS)):
        for cell in row_cells:
            if cell.value is not None:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(excel_path)
    return excel_path


def append_to_excel(content: str, excel_path: str, is_final: bool = False) -> str:
    wb = load_workbook(excel_path)
    ws = wb.active

    lines = [line.rstrip() for line in content.strip().splitlines() if line.strip()]
    current_row = ws.max_row + 1
    table_header, table_rows = parse_table(lines)

    if table_header and table_rows:
        next_id = get_next_requirement_id(ws)
        id_col = None
        if "요구사항ID" in table_header:
            id_col = table_header.index("요구사항ID")

        for row_values in table_rows:
            if id_col is not None and id_col < len(row_values):
                row_values[id_col], next_id = normalize_request_id(row_values[id_col], next_id)

            for col, value in enumerate(row_values, start=1):
                cell = ws.cell(row=current_row, column=col)
                cell.value = value
                cell.alignment = STYLE_LEFT
                cell.border = STYLE_CELL_BORDER
            current_row += 1
    elif is_final:
        for line in lines:
            cell = ws.cell(row=current_row, column=1)
            cell.value = "요구사항"
            cell.font = Font(bold=True, size=11)
            cell.fill = STYLE_SUMMARY_FILL
            cell.alignment = STYLE_LEFT
            cell.border = STYLE_CELL_BORDER

            content_cell = ws.cell(row=current_row, column=2)
            content_cell.value = line
            content_cell.alignment = STYLE_LEFT
            content_cell.border = STYLE_CELL_BORDER
            current_row += 1
    else:
        paragraph = "\n".join(lines)
        cell = ws.cell(row=current_row, column=1)
        cell.value = "청크 분석"
        cell.font = Font(bold=True, size=11)
        cell.fill = STYLE_CHUNK_FILL
        cell.alignment = STYLE_LEFT
        cell.border = STYLE_CELL_BORDER

        content_cell = ws.cell(row=current_row, column=2)
        content_cell.value = paragraph
        content_cell.alignment = STYLE_LEFT
        content_cell.border = STYLE_CELL_BORDER

    wb.save(excel_path)
    return excel_path
