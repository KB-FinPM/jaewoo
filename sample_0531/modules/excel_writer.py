from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List, Optional
import unicodedata

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

from modules.schemas import RequirementAtom, WBSItem


def _resolve_path(path: str) -> str:
    """Resolve paths even when Korean filenames are stored as NFD on macOS ZIPs."""
    p = Path(path)
    if p.exists():
        return str(p)

    parent = p.parent if str(p.parent) != '' else Path('.')
    target = unicodedata.normalize('NFC', p.name)
    if parent.exists():
        for candidate in parent.iterdir():
            if unicodedata.normalize('NFC', candidate.name) == target:
                return str(candidate)
    raise FileNotFoundError(f'템플릿 파일을 찾을 수 없습니다: {path}')


def _replace_placeholders(ws, values: Dict[str, str]):
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell) or not isinstance(cell.value, str):
                continue
            text = cell.value
            for key, value in values.items():
                text = text.replace(key, value or '')
            cell.value = text


def _header_map(ws, header_row: int = 1) -> Dict[str, int]:
    headers: Dict[str, int] = {}
    for cell in ws[header_row]:
        if cell.value is None:
            continue
        key = str(cell.value).replace('\n', '').strip()
        if key and key not in headers:
            headers[key] = cell.column
    return headers


def _clear_rows(ws, start_row: int, columns: Iterable[int]):
    for row in range(start_row, ws.max_row + 1):
        for col in columns:
            ws.cell(row=row, column=col).value = None


def _find_col(headers: Dict[str, int], names: Iterable[str], default: Optional[int] = None) -> int:
    normalized_headers = {name.replace('\n', '').strip(): col for name, col in headers.items()}
    for name in names:
        key = name.replace('\n', '').strip()
        if key in normalized_headers:
            return normalized_headers[key]
    if default is not None:
        return default
    raise KeyError(f'템플릿에서 컬럼을 찾을 수 없습니다: {list(names)}')


def save_requirement_excel(
    atoms: List[RequirementAtom],
    template_path: str,
    output_path: str,
    project_name: str,
    author: str,
):
    wb = load_workbook(_resolve_path(template_path))
    today = datetime.today().strftime('%Y-%m-%d')

    placeholder_values = {
        '{프로젝트명}': project_name,
        '{작성자명}': author,
        '{작성자}': author,
        '{작성일}': today,
    }

    if '표지' in wb.sheetnames:
        _replace_placeholders(wb['표지'], placeholder_values)

    if '개정이력' in wb.sheetnames:
        _replace_placeholders(wb['개정이력'], placeholder_values)

    ws = wb['요구사항명세서']
    headers = _header_map(ws, header_row=1)

    col_category = _find_col(headers, ['구분'], default=2)
    col_req_id = _find_col(headers, ['요구사항ID'], default=9)
    col_req_name = _find_col(headers, ['요구사항명'], default=10)
    col_req_type = _find_col(headers, ['기능/비기능요구사항'], default=11)
    col_review = _find_col(headers, ['검토의견'], default=16)

    target_columns = [col_category, col_req_id, col_req_name, col_req_type, col_review]
    start_row = 2
    _clear_rows(ws, start_row, target_columns)

    for idx, atom in enumerate(atoms, start=start_row):
        ws.cell(row=idx, column=col_category).value = atom.category
        ws.cell(row=idx, column=col_req_id).value = atom.requirement_id
        ws.cell(row=idx, column=col_req_name).value = atom.requirement_name
        ws.cell(row=idx, column=col_req_type).value = atom.requirement_type
        ws.cell(row=idx, column=col_review).value = atom.note or atom.description

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def save_wbs_excel(
    items: List[WBSItem],
    template_path: str,
    output_path: str,
):
    wb = load_workbook(_resolve_path(template_path))
    ws = wb['WBS']
    headers = _header_map(ws, header_row=1)

    col_no = headers.get('NO')
    col_level = _find_col(headers, ['레벨'], default=2)
    col_wbs_name = _find_col(headers, ['WBS명'], default=4)
    col_deliverable = _find_col(headers, ['산출물'], default=8)

    target_columns = [col for col in [col_no, col_level, col_wbs_name, col_deliverable] if col]
    start_row = 2
    _clear_rows(ws, start_row, target_columns)

    for idx, item in enumerate(items, start=start_row):
        if col_no:
            ws.cell(row=idx, column=col_no).value = idx - start_row + 1
        ws.cell(row=idx, column=col_level).value = item.level
        ws.cell(row=idx, column=col_wbs_name).value = item.wbs_name
        ws.cell(row=idx, column=col_deliverable).value = item.deliverable

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
