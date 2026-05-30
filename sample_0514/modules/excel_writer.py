from typing import List, Dict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from modules.schemas import RequirementAtom, WBSItem


def style_sheet(ws, widths: Dict[str, int]):
    header_fill = PatternFill('solid', fgColor='D9EAF7')
    header_font = Font(bold=True)
    thin = Side(style='thin', color='D9D9D9')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical='top', wrap_text=True)
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def save_requirement_excel(atoms: List[RequirementAtom], output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = '요구사항명세서'
    ws.append(['구분', '요구사항ID', '요구사항명', '기능/비기능요구사항', '비고'])
    for atom in atoms:
        ws.append([atom.category, atom.requirement_id, atom.requirement_name, atom.requirement_type, atom.note or atom.description])
    style_sheet(ws, {'A': 16, 'B': 18, 'C': 36, 'D': 22, 'E': 60})
    wb.save(output_path)


def save_wbs_excel(items: List[WBSItem], output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = 'WBS'
    ws.append(['레벨', 'WBS명', '시작예정일', '종료예정일', '작업자', '산출물'])
    for item in items:
        ws.append([item.level, item.wbs_name, item.start_date, item.end_date, item.assignee, item.deliverable])
    style_sheet(ws, {'A': 12, 'B': 48, 'C': 16, 'D': 16, 'E': 18, 'F': 40})
    wb.save(output_path)
