from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

from app.core.mapper_loader import build_context, build_placeholder_values, get_value, resolve_path
from app.schemas.pm_artifacts import RequirementAtom, WBSItem


def _replace_placeholders(ws, values: Dict[str, str]):
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell) or not isinstance(cell.value, str):
                continue
            text = cell.value
            for key, value in values.items():
                text = text.replace(key, value or '')
            cell.value = text


def _header_map(ws, header_row: int = 1) -> Dict[str, int]:
    headers: Dict[str, int] = {}
    for cell in ws[header_row]:
        if cell.value is None:
            continue
        key = str(cell.value).replace('\n', '').strip()
        if key and key not in headers:
            headers[key] = cell.column
    return headers


def _clear_rows(ws, start_row: int, columns: Iterable[int]):
    for row in range(start_row, ws.max_row + 1):
        for col in columns:
            ws.cell(row=row, column=col).value = None


def _find_col(headers: Dict[str, int], names: Iterable[str], default: Optional[int] = None, optional: bool = False) -> Optional[int]:
    normalized_headers = {name.replace('\n', '').strip(): col for name, col in headers.items()}
    for name in names or []:
        key = str(name).replace('\n', '').strip()
        if key in normalized_headers:
            return normalized_headers[key]
    if default is not None:
        return default
    if optional:
        return None
    raise KeyError(f'템플릿에서 컬럼을 찾을 수 없습니다: {list(names or [])}')


def _column_no(headers: Dict[str, int], column_mapper: Dict[str, Any]) -> Optional[int]:
    if column_mapper.get('column') is not None:
        return int(column_mapper['column'])
    return _find_col(
        headers,
        column_mapper.get('header_names', []),
        default=column_mapper.get('default_column'),
        optional=bool(column_mapper.get('optional', False)),
    )


def _apply_placeholder_sheets(wb, placeholder_sheets: List[Dict[str, Any]], context: Dict[str, str]):
    for sheet_mapper in placeholder_sheets:
        sheet_name = sheet_mapper.get('sheet_name')
        if not sheet_name or sheet_name not in wb.sheetnames:
            continue
        values = build_placeholder_values(sheet_mapper.get('placeholders', {}), source=None, context=context)
        _replace_placeholders(wb[sheet_name], values)


def _write_data_sheet(wb, items: List[Any], data_mapper: Dict[str, Any], context: Dict[str, str]):
    sheet_name = data_mapper['sheet_name']
    ws = wb[sheet_name]
    headers = _header_map(ws, header_row=int(data_mapper.get('header_row', 1)))
    start_row = int(data_mapper.get('start_row', 2))
    columns = data_mapper.get('columns', [])

    resolved_columns = []
    for column_mapper in columns:
        col_no = _column_no(headers, column_mapper)
        if col_no is None:
            continue
        resolved_columns.append((col_no, column_mapper))

    if data_mapper.get('clear_existing_rows', True):
        _clear_rows(ws, start_row, [col_no for col_no, _ in resolved_columns])

    for row_offset, item in enumerate(items):
        excel_row = start_row + row_offset
        row_number = row_offset + 1
        for col_no, column_mapper in resolved_columns:
            field_expr = column_mapper.get('field', '')
            value = get_value(item, field_expr, context=context, row_number=row_number)
            ws.cell(row=excel_row, column=col_no).value = value


def save_requirement_excel(
    atoms: List[RequirementAtom],
    template_path: str,
    output_path: str,
    project_name: str,
    author: str,
    mapper: Optional[Dict[str, Any]] = None,
):
    mapper = mapper or {
        'template_path': template_path,
        'placeholder_sheets': [
            {
                'sheet_name': '표지',
                'placeholders': {'{프로젝트명}': 'project_name', '{작성자명}': 'author', '{작성자}': 'author'},
            },
            {
                'sheet_name': '개정이력',
                'placeholders': {'{작성일}': 'today', '{작성자명}': 'author', '{작성자}': 'author'},
            },
        ],
        'data_sheet': {
            'sheet_name': '요구사항명세서',
            'header_row': 1,
            'start_row': 2,
            'columns': [
                {'field': 'category', 'header_names': ['구분'], 'default_column': 2},
                {'field': 'requirement_id', 'header_names': ['요구사항ID'], 'default_column': 9},
                {'field': 'requirement_name', 'header_names': ['요구사항명'], 'default_column': 10},
                {'field': 'requirement_type', 'header_names': ['기능/비기능요구사항'], 'default_column': 11},
                {'field': 'note|description', 'header_names': ['검토의견'], 'default_column': 16},
            ],
        },
    }

    wb = load_workbook(resolve_path(mapper.get('template_path') or template_path))
    context = build_context(project_name, author)
    _apply_placeholder_sheets(wb, mapper.get('placeholder_sheets', []), context)
    _write_data_sheet(wb, atoms, mapper['data_sheet'], context)

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def save_wbs_excel(
    items: List[WBSItem],
    template_path: str,
    output_path: str,
    mapper: Optional[Dict[str, Any]] = None,
):
    mapper = mapper or {
        'template_path': template_path,
        'data_sheet': {
            'sheet_name': 'WBS',
            'header_row': 1,
            'start_row': 2,
            'columns': [
                {'field': 'row_number', 'header_names': ['NO'], 'optional': True},
                {'field': 'level', 'header_names': ['레벨'], 'default_column': 2},
                {'field': 'wbs_name', 'header_names': ['WBS명'], 'default_column': 4},
                {'field': 'deliverable', 'header_names': ['산출물'], 'default_column': 8},
            ],
        },
    }

    wb = load_workbook(resolve_path(mapper.get('template_path') or template_path))
    _write_data_sheet(wb, items, mapper['data_sheet'], context={})

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
